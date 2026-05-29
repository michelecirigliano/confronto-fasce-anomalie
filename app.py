import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, date, time, timedelta
import re
import unicodedata
from openpyxl import load_workbook


st.set_page_config(
    page_title="Confronto fasce PD e anomalie",
    layout="wide"
)

st.title("Confronto fasce PD e anomalie")
st.write(
    "L'app confronta il file Anomalie con il file Pianificazione PD, "
    "ricostruendo le timbrature da Ent./Usc. e confrontandole con le fasce "
    "di pronta disponibilità associate al nominativo."
)


# ============================================================
# FUNZIONI BASE
# ============================================================

def normalizza_testo(valore):
    if valore is None or pd.isna(valore):
        return ""

    testo = str(valore).strip().upper()

    testo = unicodedata.normalize("NFKD", testo)
    testo = "".join(c for c in testo if not unicodedata.combining(c))

    testo = re.sub(r"[^A-Z0-9 ]+", " ", testo)
    testo = re.sub(r"\s+", " ", testo).strip()

    return testo


def normalizza_nome(cognome, nome):
    return normalizza_testo(f"{cognome} {nome}")


def nominativo_match(cognome, nome, testo_pianificazione):
    """
    Verifica se Cognome + Nome del file Anomalie è contenuto
    nella Decodifica di Matricola del file Pianificazione PD.
    Gestisce piccoli scostamenti di spazi, maiuscole, punteggiatura.
    """

    cognome_norm = normalizza_testo(cognome)
    nome_norm = normalizza_testo(nome)
    testo_norm = normalizza_testo(testo_pianificazione)

    if not cognome_norm or not nome_norm or not testo_norm:
        return False

    nominativo_1 = f"{cognome_norm} {nome_norm}"
    nominativo_2 = f"{nome_norm} {cognome_norm}"

    if nominativo_1 in testo_norm:
        return True

    if nominativo_2 in testo_norm:
        return True

    # fallback: controllo che tutte le parole principali siano presenti
    parole = [p for p in nominativo_1.split() if len(p) > 1]
    return all(p in testo_norm for p in parole)


def make_unique_headers(headers):
    """
    Replica il comportamento di pandas per colonne duplicate:
    Nome, Nome.1, Nome.2...
    """
    result = []
    counts = {}

    for h in headers:
        if h is None:
            h = ""
        h = str(h).strip()

        if h == "":
            h = "Colonna"

        if h not in counts:
            counts[h] = 0
            result.append(h)
        else:
            counts[h] += 1
            result.append(f"{h}.{counts[h]}")

    return result


def leggi_excel_righe_visibili(uploaded_file, solo_visibili=True):
    """
    Legge il primo foglio Excel usando openpyxl.
    Se solo_visibili=True, ignora le righe nascoste dal filtro Excel.
    """

    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=False))

    if not rows:
        return pd.DataFrame(), 0, 0

    # Trova la prima riga plausibile come intestazione
    header_row_index = None
    for idx, row in enumerate(rows, start=1):
        values = [cell.value for cell in row]
        non_empty = [v for v in values if v is not None and str(v).strip() != ""]
        if len(non_empty) >= 3:
            header_row_index = idx
            headers = make_unique_headers(values)
            break

    if header_row_index is None:
        return pd.DataFrame(), 0, 0

    data = []
    righe_totali = 0
    righe_lette = 0

    for row_idx in range(header_row_index + 1, ws.max_row + 1):
        righe_totali += 1

        hidden = ws.row_dimensions[row_idx].hidden

        if solo_visibili and hidden:
            continue

        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]

        # salta righe completamente vuote
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        data.append(values)
        righe_lette += 1

    df = pd.DataFrame(data, columns=headers)
    return df, righe_totali, righe_lette


def parse_data(valore):
    if valore is None or pd.isna(valore):
        return None

    if isinstance(valore, datetime):
        return valore.date()

    if isinstance(valore, date):
        return valore

    try:
        return pd.to_datetime(valore, dayfirst=True).date()
    except Exception:
        return None


def parse_ora(valore):
    """
    Converte un valore Excel/stringa in un oggetto time.
    Gestisce:
    - datetime
    - time
    - frazioni Excel
    - stringhe tipo 7:36, 07:36, 20, 20.00
    """

    if valore is None or pd.isna(valore):
        return None

    if isinstance(valore, datetime):
        return valore.time().replace(second=0, microsecond=0)

    if isinstance(valore, time):
        return valore.replace(second=0, microsecond=0)

    if isinstance(valore, (int, float)):
        # Excel spesso salva gli orari come frazione di giorno
        if 0 <= valore < 1:
            minuti = int(round(valore * 24 * 60))
            h = minuti // 60
            m = minuti % 60
            return time(h % 24, m)

        # Se è 20, significa 20:00
        if 0 <= valore <= 24:
            return time(int(valore) % 24, 0)

    testo = str(valore).strip()

    if testo.lower() in ["none", "nan", ""]:
        return None

    testo = testo.replace(".", ":")

    match = re.search(r"(\d{1,2})(?::(\d{1,2}))?", testo)
    if not match:
        return None

    h = int(match.group(1))
    m = int(match.group(2)) if match.group(2) is not None else 0

    if h > 23 or m > 59:
        return None

    return time(h, m)


def unisci_data_ora(data_base, ora_base):
    if data_base is None or ora_base is None:
        return None

    return datetime.combine(data_base, ora_base)


def format_dt(dt):
    if dt is None:
        return ""
    return dt.strftime("%d/%m/%Y %H:%M")


def format_ora(dt):
    if dt is None:
        return ""
    return dt.strftime("%H:%M")


def minuti_sovrapposizione(start_a, end_a, start_b, end_b):
    start = max(start_a, start_b)
    end = min(end_a, end_b)

    if start < end:
        minuti = int((end - start).total_seconds() // 60)
        return start, end, minuti

    return None, None, 0


# ============================================================
# TIMBRATURE FILE ANOMALIE
# ============================================================

def trova_colonne_ent_usc(df):
    """
    Trova le colonne Ent./Usc. e le accoppia in ordine.
    Questo è più robusto anche se Excel ha rinominato duplicati come Usc. 5.1.
    """

    colonne = list(df.columns)

    ent_cols = []
    usc_cols = []

    for col in colonne:
        col_norm = normalizza_testo(col)

        if col_norm.startswith("ENT"):
            ent_cols.append(col)

        if col_norm.startswith("USC"):
            usc_cols.append(col)

    # Mantiene l'ordine effettivo nel file
    coppie = []
    max_len = min(len(ent_cols), len(usc_cols))

    for i in range(max_len):
        coppie.append((ent_cols[i], usc_cols[i]))

    return coppie


def ricostruisci_timbrature_da_riga(riga, data_lavoro, coppie_ent_usc):
    """
    Da una riga del file Anomalie ricostruisce gli intervalli lavorati:
    Ent. 1 - Usc. 1
    Ent. 2 - Usc. 2
    ...
    """

    intervalli = []

    for idx, (col_ent, col_usc) in enumerate(coppie_ent_usc, start=1):
        ora_ent = parse_ora(riga.get(col_ent))
        ora_usc = parse_ora(riga.get(col_usc))

        if ora_ent is None or ora_usc is None:
            continue

        inizio = unisci_data_ora(data_lavoro, ora_ent)
        fine = unisci_data_ora(data_lavoro, ora_usc)

        if inizio is None or fine is None:
            continue

        # Se l'uscita è minore o uguale all'entrata, considero passaggio di mezzanotte
        if fine <= inizio:
            fine = fine + timedelta(days=1)

        intervalli.append({
            "progressivo": idx,
            "colonna_entrata": col_ent,
            "colonna_uscita": col_usc,
            "entrata": inizio,
            "uscita": fine,
            "testo": f"{format_ora(inizio)} - {format_ora(fine)}"
        })

    return intervalli


# ============================================================
# FASCE FILE PIANIFICAZIONE PD
# ============================================================

def estrai_fasce_da_testo(testo, solo_notturna_se_presente=True):
    """
    Estrae fasce tipo:
    20-8
    20:00-08:00
    14 - 20 / 20 - 8 DIRIGENZA

    Se solo_notturna_se_presente=True e nel testo è presente una fascia notturna
    tipo 20-8, usa solo quella, per evitare falsi positivi sulle diciture composite.
    """

    if testo is None or pd.isna(testo):
        return []

    testo = str(testo).strip()

    if testo.lower() in ["none", "nan", ""]:
        return []

    pattern = r"(\d{1,2}(?::\d{1,2})?)\s*-\s*(\d{1,2}(?::\d{1,2})?)"
    matches = re.findall(pattern, testo)

    fasce = []

    for start_txt, end_txt in matches:
        start_time = parse_ora(start_txt)
        end_time = parse_ora(end_txt)

        if start_time is None or end_time is None:
            continue

        notturna = end_time <= start_time

        fasce.append({
            "testo_fascia": f"{start_txt}-{end_txt}",
            "ora_inizio": start_time,
            "ora_fine": end_time,
            "notturna": notturna
        })

    if solo_notturna_se_presente:
        fasce_notturne = [f for f in fasce if f["notturna"]]
        if fasce_notturne:
            return fasce_notturne

    return fasce


def costruisci_intervallo_pd(data_cal, fascia):
    start_dt = datetime.combine(data_cal, fascia["ora_inizio"])
    end_dt = datetime.combine(data_cal, fascia["ora_fine"])

    if end_dt <= start_dt:
        end_dt = end_dt + timedelta(days=1)

    return start_dt, end_dt


def crea_data_da_cal(anno, mese, cal):
    try:
        cal_int = int(cal)
        return date(int(anno), int(mese), cal_int)
    except Exception:
        return None


def trasforma_pianificazione_pd(df_pd, anno, mese, solo_notturna_se_presente=True):
    """
    Trasforma il file Pianificazione PD da formato largo a formato lungo:
    ogni riga diventa una singola assegnazione PD.
    """

    righe = []

    if "Cal" not in df_pd.columns:
        return pd.DataFrame()

    for _, row in df_pd.iterrows():
        giorno_cal = row.get("Cal")
        data_cal = crea_data_da_cal(anno, mese, giorno_cal)

        if data_cal is None:
            continue

        for i in range(1, 11):
            col_dec_matricola = f"Decodifica di Matricola {i}"
            col_matricola = f"Matricola {i}"
            col_dec_orario = f"Decodifica di Orario {i}"
            col_orario = f"Orario {i}"

            if col_dec_matricola not in df_pd.columns or col_dec_orario not in df_pd.columns:
                continue

            nominativo_pd = row.get(col_dec_matricola)
            fascia_testo = row.get(col_dec_orario)

            if nominativo_pd is None or pd.isna(nominativo_pd):
                continue

            if str(nominativo_pd).strip().lower() in ["none", "nan", ""]:
                continue

            fasce = estrai_fasce_da_testo(
                fascia_testo,
                solo_notturna_se_presente=solo_notturna_se_presente
            )

            for fascia in fasce:
                inizio_pd, fine_pd = costruisci_intervallo_pd(data_cal, fascia)

                righe.append({
                    "Cal": giorno_cal,
                    "Data inizio PD": inizio_pd.date(),
                    "Progressivo": i,
                    "Nominativo PD": nominativo_pd,
                    "Nominativo PD normalizzato": normalizza_testo(nominativo_pd),
                    "Matricola PD": row.get(col_matricola) if col_matricola in df_pd.columns else "",
                    "Orario PD": row.get(col_orario) if col_orario in df_pd.columns else "",
                    "Fascia PD": fascia_testo,
                    "Fascia estratta": fascia["testo_fascia"],
                    "Inizio PD": inizio_pd,
                    "Fine PD": fine_pd,
                    "PD notturna": "Sì" if fascia["notturna"] else "No"
                })

    return pd.DataFrame(righe)


# ============================================================
# GENERAZIONE REPORT
# ============================================================

def genera_report(df_anomalie, df_pd_lungo):
    coppie_ent_usc = trova_colonne_ent_usc(df_anomalie)

    risultati = []
    righe_senza_timbrature = 0

    for idx, riga in df_anomalie.iterrows():
        cognome = riga.get("Cognome")
        nome = riga.get("Nome")
        giorno = parse_data(riga.get("Giorno"))

        if giorno is None:
            continue

        intervalli_lavoro = ricostruisci_timbrature_da_riga(
            riga,
            giorno,
            coppie_ent_usc
        )

        if not intervalli_lavoro:
            righe_senza_timbrature += 1
            continue

        # Prima filtro ampio per ridurre confronti:
        # stesso nominativo e PD che può toccare quel giorno.
        possibili_pd = []

        for _, pd_row in df_pd_lungo.iterrows():
            if not nominativo_match(cognome, nome, pd_row.get("Nominativo PD")):
                continue

            possibili_pd.append(pd_row)

        for intervallo in intervalli_lavoro:
            inizio_lav = intervallo["entrata"]
            fine_lav = intervallo["uscita"]

            for pd_row in possibili_pd:
                inizio_pd = pd_row.get("Inizio PD")
                fine_pd = pd_row.get("Fine PD")

                if inizio_pd is None or fine_pd is None:
                    continue

                start_overlap, end_overlap, minuti = minuti_sovrapposizione(
                    inizio_lav,
                    fine_lav,
                    inizio_pd,
                    fine_pd
                )

                if minuti > 0:
                    risultati.append({
                        "Cognome": cognome,
                        "Nome": nome,
                        "Nominativo Anomalie": f"{cognome} {nome}",
                        "Data timbratura": giorno.strftime("%d/%m/%Y"),
                        "Entrata": format_dt(inizio_lav),
                        "Uscita": format_dt(fine_lav),
                        "Intervallo timbratura": f"{format_dt(inizio_lav)} - {format_dt(fine_lav)}",
                        "Cal pianificazione": pd_row.get("Cal"),
                        "Nominativo PD trovato": pd_row.get("Nominativo PD"),
                        "Orario PD": pd_row.get("Orario PD"),
                        "Fascia PD": pd_row.get("Fascia PD"),
                        "Fascia estratta": pd_row.get("Fascia estratta"),
                        "Inizio PD": format_dt(inizio_pd),
                        "Fine PD": format_dt(fine_pd),
                        "Inizio sovrapposizione": format_dt(start_overlap),
                        "Fine sovrapposizione": format_dt(end_overlap),
                        "Minuti sovrapposti": minuti,
                        "Ore sovrapposte": round(minuti / 60, 2),
                        "Tipo anomalia": "Sovrapposizione tra timbratura e fascia PD",
                        "Nota": (
                            "Il soggetto risulta timbrato in un intervallo che si sovrappone "
                            "alla fascia di pronta disponibilità pianificata."
                        )
                    })

    report = pd.DataFrame(risultati)

    return report, coppie_ent_usc, righe_senza_timbrature


def crea_file_excel_report(report, df_pd_lungo, coppie_ent_usc):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        report.to_excel(writer, index=False, sheet_name="Anomalie rilevate")

        df_pd_export = df_pd_lungo.copy()
        for col in ["Inizio PD", "Fine PD"]:
            if col in df_pd_export.columns:
                df_pd_export[col] = df_pd_export[col].apply(format_dt)

        df_pd_export.to_excel(writer, index=False, sheet_name="PD trasformata")

        df_coppie = pd.DataFrame(
            [{"Entrata": e, "Uscita": u} for e, u in coppie_ent_usc]
        )
        df_coppie.to_excel(writer, index=False, sheet_name="Colonne Ent-Usc usate")

        workbook = writer.book
        header_format = workbook.add_format({
            "bold": True,
            "bg_color": "#D9EAF7",
            "border": 1
        })
        cell_format = workbook.add_format({
            "border": 1
        })

        for sheet_name, worksheet in writer.sheets.items():
            worksheet.freeze_panes(1, 0)

            if sheet_name == "Anomalie rilevate":
                df_ref = report
            elif sheet_name == "PD trasformata":
                df_ref = df_pd_export
            else:
                df_ref = df_coppie

            for col_num, value in enumerate(df_ref.columns):
                worksheet.write(0, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 24)

            for row_num in range(1, len(df_ref) + 1):
                for col_num in range(len(df_ref.columns)):
                    worksheet.write(row_num, col_num, df_ref.iloc[row_num - 1, col_num], cell_format)

    output.seek(0)
    return output


# ============================================================
# INTERFACCIA STREAMLIT
# ============================================================

st.subheader("1. Caricamento file")

file_anomalie = st.file_uploader(
    "Carica il file Anomalie",
    type=["xlsx"],
    key="file_anomalie"
)

file_pd = st.file_uploader(
    "Carica il file Pianificazione PD",
    type=["xlsx"],
    key="file_pd"
)

st.subheader("2. Parametri")

col_a, col_b, col_c = st.columns(3)

with col_a:
    mese = st.selectbox(
        "Mese della pianificazione PD",
        options=list(range(1, 13)),
        index=3,
        format_func=lambda x: {
            1: "Gennaio",
            2: "Febbraio",
            3: "Marzo",
            4: "Aprile",
            5: "Maggio",
            6: "Giugno",
            7: "Luglio",
            8: "Agosto",
            9: "Settembre",
            10: "Ottobre",
            11: "Novembre",
            12: "Dicembre"
        }[x]
    )

with col_b:
    anno = st.number_input(
        "Anno della pianificazione PD",
        min_value=2020,
        max_value=2035,
        value=2026,
        step=1
    )

with col_c:
    solo_righe_visibili = st.checkbox(
        "Leggi solo righe visibili/filtrate del file Anomalie",
        value=True
    )

solo_notturna = st.checkbox(
    "Se una fascia contiene una parte notturna tipo 20-8, considera solo quella",
    value=True,
    help=(
        "Utile quando la decodifica contiene diciture composite tipo "
        "'8-20 / 20-8 DIRIGENZA' ma l'anomalia da cercare riguarda la parte 20-8."
    )
)

if file_anomalie is not None and file_pd is not None:

    st.subheader("3. Lettura file")

    df_anomalie, righe_totali, righe_lette = leggi_excel_righe_visibili(
        file_anomalie,
        solo_visibili=solo_righe_visibili
    )

    # Per la pianificazione PD di solito non serve rispettare filtri
    file_pd.seek(0)
    df_pd = pd.read_excel(file_pd)
    righe_pd_lette = len(df_pd)

    st.write(f"Righe dati file Anomalie lette: **{righe_lette}** su **{righe_totali}**")
    st.write(f"Righe dati file Pianificazione PD lette: **{righe_pd_lette}**")

    if df_anomalie.empty:
        st.error("Il file Anomalie risulta vuoto o non leggibile.")
        st.stop()

    if df_pd.empty:
        st.error("Il file Pianificazione PD risulta vuoto o non leggibile.")
        st.stop()

    st.subheader("4. Anteprima file")

    col1, col2 = st.columns(2)

    with col1:
        st.write("Anteprima file Anomalie")
        st.dataframe(df_anomalie.head(20), use_container_width=True)

    with col2:
        st.write("Anteprima file Pianificazione PD")
        st.dataframe(df_pd.head(20), use_container_width=True)

    st.subheader("5. Controllo colonne obbligatorie")

    colonne_obbligatorie_anomalie = ["Cognome", "Nome", "Giorno"]
    mancanti_anomalie = [c for c in colonne_obbligatorie_anomalie if c not in df_anomalie.columns]

    if mancanti_anomalie:
        st.error(f"Nel file Anomalie mancano queste colonne obbligatorie: {mancanti_anomalie}")
        st.stop()

    if "Cal" not in df_pd.columns:
        st.error("Nel file Pianificazione PD manca la colonna obbligatoria 'Cal'.")
        st.stop()

    coppie_preview = trova_colonne_ent_usc(df_anomalie)

    if not coppie_preview:
        st.error("Non sono state trovate coppie di colonne Ent./Usc. nel file Anomalie.")
        st.stop()

    st.write("Coppie Ent./Usc. rilevate nel file Anomalie:")
    st.dataframe(
        pd.DataFrame([{"Entrata": e, "Uscita": u} for e, u in coppie_preview]),
        use_container_width=True
    )

    if st.button("Avvia confronto"):

        with st.spinner("Elaborazione in corso..."):
            df_pd_lungo = trasforma_pianificazione_pd(
                df_pd,
                anno=int(anno),
                mese=int(mese),
                solo_notturna_se_presente=solo_notturna
            )

            if df_pd_lungo.empty:
                st.error(
                    "La pianificazione PD trasformata è vuota. "
                    "Controlla che siano presenti le colonne Decodifica di Matricola 1-10 "
                    "e Decodifica di Orario 1-10."
                )
                st.stop()

            report, coppie_ent_usc, righe_senza_timbrature = genera_report(
                df_anomalie,
                df_pd_lungo
            )

        st.subheader("6. Pianificazione PD trasformata")

        df_pd_preview = df_pd_lungo.copy()
        df_pd_preview["Inizio PD"] = df_pd_preview["Inizio PD"].apply(format_dt)
        df_pd_preview["Fine PD"] = df_pd_preview["Fine PD"].apply(format_dt)

        st.dataframe(df_pd_preview, use_container_width=True)

        st.subheader("7. Report anomalie")

        st.write(f"Righe Anomalie senza timbrature Ent./Usc. interpretabili: **{righe_senza_timbrature}**")

        if report.empty:
            st.success("Nessuna sovrapposizione trovata con i criteri attuali.")
        else:
            st.warning(f"Sono state trovate **{len(report)}** sovrapposizioni/anomalie.")
            st.dataframe(report, use_container_width=True)

            excel_report = crea_file_excel_report(report, df_pd_lungo, coppie_ent_usc)

            st.download_button(
                label="Scarica report anomalie in Excel",
                data=excel_report,
                file_name="report_anomalie_pd.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

else:
    st.info("Carica entrambi i file Excel per avviare il confronto.")
